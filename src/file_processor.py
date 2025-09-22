# -*- coding: utf-8 -*-
"""
Created on Tue Nov  8 22:28:15 2022

@author: XIAOQING CHEN
"""

import pandas as pd
# import numpy as np
import xlwings as xw
import os
# import re
import time
# import datetime
# import shutil
# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
# from openpyxl.styles import colors
# from openpyxl.styles import Alignment
# from openpyxl.styles import PatternFill
# from openpyxl.styles import Side,Border

#%%
class LocalFolder():
    def __init__(self, folder_path):
        """       
        Parameters
        ------------------------
        folder_path: Target Local Folder path
        ------------------------
        """
        self.sub_folder = []
        self.file_list = []
        self.file_location = {}
        self.system_date = time.strftime("%Y%m%d",time.localtime(time.time())) #系统时间YYYYMMMDD(字符串)

        if os.path.exists(folder_path) != False:
        #self.sub_folder, self.file_list, self.file_location
            for root,dirs,files in os.walk(folder_path):
                self.sub_folder = self.sub_folder + dirs
                self.file_list = self.file_list + files
                for file in files:
                    self.file_location[file] = root
        else:
            print('Please note: The provided local folder path does not exist. \n')

    def NewFolder(self, folder_path, *folder_name:str):
        """  
        Purpose
        ------------------------
        Create new folder in Target local path
        ------------------------

        Parameters
        ------------------------
        folder_path: Target Local Folder path
        folder_name: New folder name, such as 'Raw file', 'system_date\Raw file' #创建一级文件夹system_date及二级文件夹Raw file
        ------------------------
        """
        if os.path.exists(folder_path) != False:

            for f in folder_name:
                new_path = os.path.join(folder_path,f)
                if not os.path.exists(new_path):
                    os.makedirs(new_path) #创建新的文件夹
                else:
                    print(f'File path <{new_path}> exist.\n')

        return 'New Folder complete'

    def find_file_name(self,file_types:list, file_key:tuple):
        """       
        Parameters
        ------------------------
        file_types: Types of Target file, such as: .xls, .xlsx, .csv, .pdf. etc.
        file_key: Keys of Target file
        ------------------------

        Attributes
        ------------------------
        self.target_file_list: List of file names containing the provided keywords
        self.target_file: Filename of the last file
        self.target_file_path: File Path of the last file
        self.target_file_ext: File extension name of the last file
        ------------------------
        """
        self.target_file_list = [] #含提供关键词的所有文件的文件名列表
        self.target_file = False
        self.target_file_path = False
        self.target_file_ext = False
        
        if len(self.file_list) > 0:
        
            #self.target_file_list
            for file in self.file_list:
                file_remain = True
                file_name,file_ext = os.path.splitext(file)
                
                if file_ext.lower() not in file_types:
                    file_remain = False
                    continue
                
                for key in file_key:
                    if key.lower() not in file_name.lower():
                        file_remain = False
                        break
    
                if file_remain == True:
                    self.target_file_list.append(file)
            
            #self.target_file, self.target_file_path, self.target_file_ext
            if len(self.target_file_list) != 0:
                self.target_file = sorted(self.target_file_list)[-1]
                self.target_file_path = self.file_location[self.target_file]
                file_name, self.target_file_ext = os.path.splitext(self.target_file)

            if len(self.target_file_list) == 0:
                print(f'Please note: Target file contains key {file_key} not found.\n')
            elif len(self.target_file_list) > 1:
                print(f'Please note: More than one files contain keys {file_key}.\n')
        
        return 'Find file name complete'

    def read_excel(self,*file_key:str, tab, tab_exact = True, header):
        self.find_file_name(['xls','xlsx','xlsm','csv'], file_key)
        
        df = pd.DataFrame()
        full_file = os.path.join(self.target_file_path, self.target_file)
        
        try:
            df = pd.read_excel(full_file, sheet_name = tab, header = header)
        except:
            if self.target_file.split('.')[-1] == 'csv':
                df = pd.read_csv(full_file, header = header)
            else:
                
                try:
                    xl = pd.ExcelFile(full_file)

                    actual_tab_name = None
                    for tab_name in xl.sheet_names:
                        if tab.lower() in tab_name.lower():
                            actual_tab_name = tab_name
                            break
                    
                    if actual_tab_name == None:
                        print(f'Tab name {tab} not found in target file.\n')
                        
                    df = pd.read_excel(full_file, sheet_name = actual_tab_name, header = header)

                except:
                    wb = xw.Book(full_file)
                    
                    if type(tab) == str:
                        wb_num = len(wb.sheets)
                        actual_tab = None
                        for i in range(0,wb_num):
                            sheet = wb.sheets[i]
                            if tab.lower() in sheet.name.lower():
                                actual_tab = i
                                rng = sheet.used_range
                                nrow = rng.rows.count
                                ncol = rng.columns.count
                                ncolL = get_column_letter(ncol)
                                df = pd.DataFrame(sheet.range(f'A1:{ncolL}{nrow}').value)
                                break
                            
                        if actual_tab  == None:
                            print(f'Tab name {tab} not found in target file.\n')
                        
                    else:
                        actual_tab_index = 0
                        if tab == None:
                            actual_tab_index = 0
                        else:
                            actual_tab_index = tab
                        
                        sheet = wb.sheets[actual_tab_index]
                        rng = sheet.used_range
                        nrow = rng.rows.count
                        ncol = rng.columns.count
                        ncolL = get_column_letter(ncol)
                        df = pd.DataFrame(sheet.range(f'A1:{ncolL}{nrow}').value)

                    wb.close()
                    wb.app.quit()
        
        print(f'File {self.target_file} {tab} read complete.\n')
        
        return df

    def output_excel(self, folder_path, output_file_name:str, *args:(pd.DataFrame,str,bool,bool)):
        """       
        Purpose
        ------------------------
        Output dataframe as excel.
        ------------------------
        
        Parameters
        ------------------------
        output_file_name: Name of Output Excel.
        args: Dataframe, Name of Output Tab, if header be output, if index be output
        ------------------------
        """
        if os.path.exists(folder_path) != False:
            
            writer = pd.ExcelWriter(os.path.join(folder_path,output_file_name))
            for arg in args:
                arg[0].to_excel(writer, sheet_name = arg[1], header = arg[2], index = arg[3])
            
            writer.save()
